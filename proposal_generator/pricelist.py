"""
Manufacturer price list: download / parse, then apply it to a configuration.

A price list is a mapping of SKU -> (description, manufacturer, unit list price).
We support three sources:

  * a local CSV          (columns: sku, description, manufacturer, unit_price[, category])
  * a local XLSX         (same columns; first sheet, header row)
  * a URL pointing to either of the above

Real manufacturer price lists are partner-gated, so a representative sample is
bundled (sample_data/pricelist_sample.csv) and used when no source is given.
"""

from __future__ import annotations

import csv
import io
import os
from dataclasses import dataclass
from typing import Dict, List, Tuple
from urllib.parse import urlparse

from .models import LineItemSpec, PricedLineItem


# ---------------------------------------------------------------------------
# Price list container
# ---------------------------------------------------------------------------

@dataclass
class PriceRecord:
    sku: str
    description: str
    manufacturer: str
    unit_price: float
    category: str = "other"


class PriceList:
    def __init__(self, records: Dict[str, PriceRecord] | None = None):
        self._records: Dict[str, PriceRecord] = records or {}

    def __len__(self) -> int:
        return len(self._records)

    def __contains__(self, sku: str) -> bool:
        return sku.upper() in self._records

    def get(self, sku: str) -> PriceRecord | None:
        return self._records.get(sku.upper())

    def add(self, rec: PriceRecord) -> None:
        self._records[rec.sku.upper()] = rec

    # -- parsing -----------------------------------------------------------

    @classmethod
    def from_rows(cls, rows: List[dict]) -> "PriceList":
        pl = cls()
        for row in rows:
            # tolerate header variations / casing / whitespace
            norm = {(k or "").strip().lower(): (v or "") for k, v in row.items()}
            sku = str(norm.get("sku") or norm.get("part") or norm.get("part_number") or "").strip()
            if not sku:
                continue
            price_raw = str(
                norm.get("unit_price") or norm.get("price")
                or norm.get("list_price") or norm.get("msrp") or "0"
            )
            price = _parse_price(price_raw)
            pl.add(PriceRecord(
                sku=sku,
                description=str(norm.get("description") or norm.get("desc") or "").strip(),
                manufacturer=str(norm.get("manufacturer") or norm.get("mfr") or norm.get("vendor") or "").strip(),
                unit_price=price,
                category=str(norm.get("category") or "other").strip().lower() or "other",
            ))
        return pl

    @classmethod
    def from_csv_text(cls, text: str) -> "PriceList":
        reader = csv.DictReader(io.StringIO(text))
        return cls.from_rows(list(reader))

    @classmethod
    def from_csv_file(cls, path: str) -> "PriceList":
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            return cls.from_csv_text(f.read())

    @classmethod
    def from_xlsx_file(cls, path: str) -> "PriceList":
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as e:
            raise SystemExit(
                "openpyxl is required to read .xlsx price lists "
                "(`pip install openpyxl`), or supply a CSV."
            ) from e
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(c).strip().lower() if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            return cls()
        rows = []
        for r in rows_iter:
            rows.append({header[i]: r[i] for i in range(min(len(header), len(r)))})
        return cls.from_rows(rows)

    @classmethod
    def from_bytes(cls, data: bytes, hint: str = "") -> "PriceList":
        if hint.lower().endswith((".xlsx", ".xlsm")):
            import tempfile
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(data)
                tmp_path = tmp.name
            try:
                return cls.from_xlsx_file(tmp_path)
            finally:
                os.unlink(tmp_path)
        return cls.from_csv_text(data.decode("utf-8-sig", errors="replace"))


# ---------------------------------------------------------------------------
# Source resolution (URL / path / sample)
# ---------------------------------------------------------------------------

def _parse_price(s: str) -> float:
    cleaned = str(s).replace("$", "").replace(",", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _sample_path() -> str:
    here = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(here, "..", "sample_data", "pricelist_sample.csv")


def load_price_list(source: str = "") -> Tuple[PriceList, str]:
    """
    Resolve a price-list source into a PriceList.

    Returns (price_list, description_of_source).
    """
    if not source:
        path = _sample_path()
        return PriceList.from_csv_file(path), f"bundled sample ({os.path.basename(path)})"

    parsed = urlparse(source)
    if parsed.scheme in ("http", "https"):
        try:
            import requests  # already a project dependency
        except ImportError as e:
            raise SystemExit("requests is required to download a price list.") from e
        print(f"[PRICELIST] Downloading {source} ...")
        r = requests.get(source, timeout=60)
        r.raise_for_status()
        pl = PriceList.from_bytes(r.content, hint=parsed.path)
        return pl, source

    # local file
    if source.lower().endswith((".xlsx", ".xlsm")):
        return PriceList.from_xlsx_file(source), source
    return PriceList.from_csv_file(source), source


# ---------------------------------------------------------------------------
# Apply a price list to a configuration
# ---------------------------------------------------------------------------

@dataclass
class QuoteTotals:
    subtotal_list: float
    subtotal_net: float
    total_savings: float
    discount_avg_pct: float


def apply_to_config(
    price_list: PriceList,
    line_items: List[LineItemSpec],
    default_discount_pct: float = 0.0,
) -> Tuple[List[PricedLineItem], QuoteTotals, List[str]]:
    """
    Join requested SKUs to the price list. Returns priced line items, totals,
    and a list of warnings for any SKUs not found in the price list.
    """
    priced: List[PricedLineItem] = []
    warnings: List[str] = []

    for li in line_items:
        rec = price_list.get(li.sku)
        if rec is None:
            warnings.append(f"SKU not found in price list: {li.sku} (qty {li.qty}) — priced at $0")
            rec = PriceRecord(
                sku=li.sku,
                description=li.description_override or "(SKU not found in price list)",
                manufacturer="",
                unit_price=0.0,
                category=li.category,
            )
        discount = li.discount_pct if li.discount_pct is not None else default_discount_pct
        priced.append(PricedLineItem(
            sku=li.sku,
            description=li.description_override or rec.description,
            manufacturer=rec.manufacturer,
            category=li.category if li.category != "other" else rec.category,
            qty=li.qty,
            unit_list=rec.unit_price,
            discount_pct=discount,
        ))

    subtotal_list = round(sum(p.extended_list for p in priced), 2)
    subtotal_net = round(sum(p.extended for p in priced), 2)
    savings = round(subtotal_list - subtotal_net, 2)
    avg_disc = round((savings / subtotal_list * 100.0), 1) if subtotal_list else 0.0

    totals = QuoteTotals(
        subtotal_list=subtotal_list,
        subtotal_net=subtotal_net,
        total_savings=savings,
        discount_avg_pct=avg_disc,
    )
    return priced, totals, warnings
